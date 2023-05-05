#Advanced Guide To Python 3 Programming by John Hunt Chapter 20 Working With CSV Files
import csv
with open("csvfilebeatles.csv", "w", newline="") as variablewritecsvfile:
    writer = csv.writer(variablewritecsvfile)
    writer.writerow(["She Loves You", "Sep 1963"])
    writer.writerow(["I Want To Hold Your Hand", "Dec 1963"])
    writer.writerow(["Can't Buy Me Love", "Apr 1964"])
    writer.writerow(["A Hard Days Night", "July 1964"])
with open("csvfilebeatles.csv", "r", newline="") as variablereadcsvfile:
    reader = csv.reader(variablereadcsvfile)
    for eachreader in reader:
        print(*eachreader, sep=", ")
        '''
        She Loves You, Sep 1963
        I Want To Hold Your Hand, Dec 1963
        Can't Buy Me Love, Apr 1964
        A Hard Days Night, July 1964
        '''
with open("csvfilebeatles.csv", "r", newline="") as variablereadcsvfile:
    reader = csv.reader(variablereadcsvfile)
    for eachreader in reader:
        print(eachreader, sep=", ")
        '''
        ['She Loves You', 'Sep 1963']
        ['I Want To Hold Your Hand', 'Dec 1963']
        ["Can't Buy Me Love", 'Apr 1964']
        ['A Hard Days Night', 'July 1964']
        '''
with open("csvdictionaryheaders.csv", "w", newline="") as variablewritecsvfile:
    columnheaders = ["firstname", "lastname", "resultnumber"]
    writer = csv.DictWriter(variablewritecsvfile, fieldnames=columnheaders)
    writer.writeheader()
    writer.writerow({"firstname": "John", "lastname": "Smith", "resultnumber": 54})
    writer.writerow({"firstname": "Jane", "lastname": "Lewis", "resultnumber": 63})
    writer.writerow({"firstname": "Chris", "lastname": "Davies", "resultnumber": 72})
with open("csvdictionaryheaders.csv", "r", newline="") as variablereadcsvfile:
    reader = csv.DictReader(variablereadcsvfile)
    for heading in reader.fieldnames:
        print(heading, end=" ")
    print("\b")
    for eachreader in reader:
        print(eachreader["firstname"], eachreader["lastname"], eachreader["resultnumber"])
    '''
    firstname lastname resultnumber 
    John Smith 54
    Jane Lewis 63
    Chris Davies 72
    '''

#Advanced Guide To Python 3 Programming by John Hunt Chapter 21 Working With Excel Files
from openpyxl import Workbook #import the OpenPyXL library using the Workbook class
wb = Workbook() #create an instance Workbook saved to variable wb
ws = wb.active #ws is worksheet
ws = wb.create_sheet("Create worksheet name here")
ws.title = "Worksheet title"
ws.sheet_properties.tabColor = "1072BA" #background tab color default is white.  Change the tab color providing an RRGGBB color code.
ws["A1"] = "Input at cell A1"
print(ws["A1"]) #print <Cell 'Worksheet title'.A1>
print(ws["A1"].value) #print Input at cell A1
cell = ws["A1"]
print(cell.value) #print Input at cell A1
cellb4 = ws.cell(row=4, column=2, value=10)
print(cellb4.value) #print 10
ws.append([1, 2, 3])
wb.save("tempexcel.xlsx")
#RM:  stopped reading chapter